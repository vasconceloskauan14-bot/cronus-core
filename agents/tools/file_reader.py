"""
File Reader Tool — ULTIMATE CRONUS
Lê PDF, CSV, DOCX, JSON, TXT, Markdown e retorna conteúdo como texto.
"""

import csv
import io
import json
import os
from pathlib import Path
from .base_tool import BaseTool, ToolResult


class FileReaderTool(BaseTool):
    """
    Lê arquivos de vários formatos e retorna conteúdo como texto.
    Suporta: PDF, CSV, DOCX, JSON, TXT, MD, XLSX.
    """

    name = "file_reader"
    description = "Lê arquivos (PDF, CSV, DOCX, JSON, TXT) e retorna o conteúdo. Útil para analisar documentos."

    def __init__(self, max_chars: int = 50_000):
        self.max_chars = max_chars

    def run(self, path: str, sheet: str = "") -> ToolResult:
        """
        Lê um arquivo.
        Args:
            path: Caminho do arquivo
            sheet: Nome da aba (apenas para XLSX)
        """
        p = Path(path)
        if not p.exists():
            return ToolResult(success=False, output="", error=f"Arquivo não encontrado: {path}")

        ext = p.suffix.lower()
        try:
            if ext == ".pdf":
                return self._read_pdf(p)
            elif ext == ".csv":
                return self._read_csv(p)
            elif ext in (".xlsx", ".xls"):
                return self._read_excel(p, sheet)
            elif ext in (".docx", ".doc"):
                return self._read_docx(p)
            elif ext == ".json":
                return self._read_json(p)
            elif ext in (".txt", ".md", ".py", ".js", ".ts", ".html", ".xml", ".yaml", ".yml"):
                return self._read_text(p)
            else:
                return self._read_text(p)  # tenta como texto
        except Exception as e:
            return ToolResult(success=False, output="", error=str(e))

    def _read_pdf(self, p: Path) -> ToolResult:
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(str(p)) as pdf:
                for i, page in enumerate(pdf.pages):
                    t = page.extract_text() or ""
                    text_parts.append(f"[Página {i+1}]\n{t}")
            text = "\n\n".join(text_parts)
        except ImportError:
            try:
                import PyPDF2
                text_parts = []
                with open(p, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    for i, page in enumerate(reader.pages):
                        text_parts.append(f"[Página {i+1}]\n{page.extract_text() or ''}")
                text = "\n\n".join(text_parts)
            except ImportError:
                return ToolResult(success=False, output="", error="Instale pdfplumber: pip install pdfplumber")
        return ToolResult(
            success=True,
            output=text[:self.max_chars],
            metadata={"file": str(p), "format": "pdf", "chars": len(text)},
        )

    def _read_csv(self, p: Path) -> ToolResult:
        rows = []
        with open(p, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(", ".join(row))
        text = "\n".join(rows)
        return ToolResult(
            success=True,
            output=text[:self.max_chars],
            metadata={"file": str(p), "format": "csv", "rows": len(rows)},
        )

    def _read_excel(self, p: Path, sheet: str = "") -> ToolResult:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(", ".join(str(c or "") for c in row))
            text = "\n".join(rows)
            return ToolResult(
                success=True,
                output=text[:self.max_chars],
                metadata={"file": str(p), "format": "xlsx", "sheet": ws.title, "rows": len(rows)},
            )
        except ImportError:
            return ToolResult(success=False, output="", error="Instale openpyxl: pip install openpyxl")

    def _read_docx(self, p: Path) -> ToolResult:
        try:
            import docx
            doc = docx.Document(str(p))
            text = "\n".join(para.text for para in doc.paragraphs)
            return ToolResult(
                success=True,
                output=text[:self.max_chars],
                metadata={"file": str(p), "format": "docx"},
            )
        except ImportError:
            return ToolResult(success=False, output="", error="Instale python-docx: pip install python-docx")

    def _read_json(self, p: Path) -> ToolResult:
        data = json.loads(p.read_text(encoding="utf-8"))
        text = json.dumps(data, indent=2, ensure_ascii=False)
        return ToolResult(
            success=True,
            output=text[:self.max_chars],
            metadata={"file": str(p), "format": "json"},
        )

    def _read_text(self, p: Path) -> ToolResult:
        text = p.read_text(encoding="utf-8", errors="replace")
        return ToolResult(
            success=True,
            output=text[:self.max_chars],
            metadata={"file": str(p), "format": p.suffix, "chars": len(text)},
        )

    def schema(self) -> dict:
        return {
            "name": self.name,
            "description": self.description,
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Caminho do arquivo"},
                    "sheet": {"type": "string", "description": "Aba do Excel (opcional)"},
                },
                "required": ["path"],
            },
        }
